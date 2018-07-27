#Copyright (c) 2018 Joram Puumala
#
#Permission is hereby granted, free of charge, to any person obtaining a copy
#of this software and associated documentation files (the "Software"), to deal
#in the Software without restriction, including without limitation the rights
#to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#copies of the Software, and to permit persons to whom the Software is
#furnished to do so, subject to the following conditions:
#
#The above copyright notice and this permission notice shall be included in all
#copies or substantial portions of the Software.
#
#THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#SOFTWARE.


#!/usr/bin/env python3

from getpass import getpass
import re
import sys
import subprocess
import paramiko
import os
import openpyxl
import datetime
from CoSSH.Utils.FileWriting import InPlaceReplacement
from CoSSH.Utils.Hashing import LocalHash

class SSHConfiguration():

	# initialize SSH connection
	def __init__(self, conn, sftp):
		self.conn = conn
		self.sftp = sftp

	# close SSH connection
	def close_ssh(self):
		self.conn.close()
		self.sftp.close()

	# this function writes to router's latest_update.txt
	def update_name(self, args):
		func_stat = 2
		update_stamp = args.split(",", 1)[0]

		# first check if cossh directory exists under /root
		check_initial = "ls /root/cossh/"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_initial)
		check_status = ssh_stdout.channel.recv_exit_status()

		# if the directory doesn't exist, create one
		if check_status != 0:
			create_cossh = "mkdir /root/cossh/"
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(create_cossh)
			check_status = ssh_stdout.channel.recv_exit_status()

			if check_status != 0:
				error_msg = "Failed to create update-stamp"
				return error_msg, func_stat

		# write update
		new_stamp = "echo '" + update_stamp + "' > /root/cossh/latest_update.txt"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(new_stamp)
		check_status = ssh_stdout.channel.recv_exit_status()

		if check_status == 0:
			status_msg = "Update name: " + update_stamp
			func_stat = 0
		else:
			status_msg = "Failed to create update-stamp"

		return status_msg, func_stat

	# this function reads latest update and prints it
	def latest_update(self, args):
		func_stat = 1
		check_updatefile = "cat /root/cossh/latest_update.txt"

		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_updatefile)
		check_status = ssh_stdout.channel.recv_exit_status()

		if check_status != 0:
			status_msg = "No prior updates found"
		else:
			status_msg = ssh_stdout.readlines()
			status_msg = "Latest update: " + status_msg[0].strip()
			func_stat = 0

		return status_msg, func_stat

	# this function adds a new client to a group (for group configuration)
	def add_client(self, args):
		func_stat = 2
		client_name = args.split(",", 1)[0]
		client_group = args.split(",", 2)[1]
		client_conf = "/etc/cossh/clients/clients.conf"
		private_key_file = "/etc/cossh/keys/cossh-key_" + client_group
		public_key_file = "/etc/cossh/keys/cossh-key_" + client_group + ".pub"
		check_ssh_dir = "ls /root/.ssh"
		serial = self.get_serial()

		if not os.path.exists(client_conf):
			open(client_conf, 'a').close()

		# read current client.conf
		with open(client_conf) as clcf:
			content = list(map(str.strip,clcf.readlines()))

			# check if client is already in a group
			# a client can only be in one group at a time
			if any(serial in s for s in content):
				status_msg = "Client '" + serial + "' is already in a group"
				func_stat = 1
				return status_msg, func_stat

			client_bool = False

			# find the group the client will join
			for line in content:
				if line.strip().startswith("@@") and line.strip().endswith("@@"):
					if line.strip() == "@@" + client_group + "@@":
						client_bool = True
					else:
						client_bool = False

				# check if new client's IPv4 address is already occupied in the group
				if client_bool == True and not line.strip().startswith("@@") and line.strip() != "":
					client_ip = line.strip().split(":", 1)[0]
					if client_name == client_ip:
						status_msg = "Client IP '" + client_name + "' already belongs to the group"
						return status_msg, func_stat

			# if group doesn't exist, mark it as a new group
			if not "@@" + client_group + "@@" in content:
				new_group = True
			else:
				new_group = False

		# check if groups SSH keys exist
		if not os.path.exists(private_key_file) and not os.path.exists(public_key_file):
			passwd_success = False
			print("Created new group '" + client_group + "'")

			try:
				# ask to set a new password for the group
				while passwd_success == False:
					key_passwd = getpass("Create password for the group: ")
					verify_passwd = getpass("Verify password: ")
					if key_passwd == verify_passwd:
						if re.match(r'[A-Za-z0-9@#$%^&+=!?-]{8,}', key_passwd):
							passwd_success = True
						else:
							print("Minimum length for password is 8 characters, allowed characters are A-Za-z0-9@#$%^&+=!?-")
			except KeyboardInterrupt:
				print("Aborted")

			except Exception as e:
				print(e)

			# create new encrypted SSH keys for the group
			create_keys = 'ssh-keygen -b 2048 -t rsa -f /etc/cossh/keys/cossh-key_' + client_group + ' -q -N "' + key_passwd + '"'
			subprocess.call([create_keys], shell=True)

		# if it's a new group, add group and its first client
		if new_group == True:
			with open(client_conf, "a") as clcf:
				clcf.write("\n@@" + client_group + "@@\n")
				clcf.write(client_name + ":" + str(serial) + "\n")

		# if group exists, add client
		else:
			#add_client = "sed -i '/@@" + client_group + "@@/a " + client_name + ":" + str(serial) + " ' " + client_conf
			#subprocess.call([add_client], shell=True)
			InPlaceReplacement.after_string("@@" + client_group + "@@", client_name + ":" + str(serial), client_conf)


		with open(public_key_file) as pubk:
			public_key = pubk.read().replace("\n", "")

		# check if .ssh dir exists in router
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_ssh_dir)
		check_status = ssh_stdout.channel.recv_exit_status()

		# if the directory doesn't exist, create one
		if check_status != 0:
			create_ssh_dir = "mkdir /root/.ssh"
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(create_ssh_dir)
			dir_status = ssh_stdout.channel.recv_exit_status()

		# add public key to router's authorized keys and create a copy under /opt directory
		# this is because during factory reset/firmware update everything will be wiped but /opt directory and configuration
		# also during startup, check whether /root/.ssh exists, if not copy from the backup
		if check_status == 0 or dir_status == 0:
			add_public_key = "echo " + public_key + " >> /root/.ssh/authorized_keys"
			copy_key = "cp -r /root/.ssh /opt/"
			add_startup_check = 'echo "if [ ! -d /root/.ssh/ ]; then cp -r /opt/.ssh/ /root/;fi" >> /etc/rc.local'
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(add_public_key)
			key_status = ssh_stdout.channel.recv_exit_status()
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(copy_key)
			copy_status = ssh_stdout.channel.recv_exit_status()
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(add_startup_check)
			startup_status = ssh_stdout.channel.recv_exit_status()

			if key_status == 0 and copy_status == 0 and startup_status == 0:
				status_msg = "Public-key succesfully installed to router, client '" + serial + "' added to group '" + client_group + "'"
				func_stat = 0
			elif key_status == 0 and copy_status != 0 or startup_status != 0:
				status_msg = "Public-key succesfully installed to router, client '" + serial + "' added to group '" + client_group + "', but failed to create backup"
				func_stat = 1
			else:
				status_msg = "Failed to install public-key to router"
				func_stat = 2
		else:
			status_msg = "Failed to create directory for authorized_keys"
			func_stat = 2

		return status_msg, func_stat
		

	# run a custom command in router
	def router_command(self, args):
		func_stat = 2
		user_command = args.split(",", 1)[0]

		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(user_command)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		if cmd_status == 0:
			func_stat = 0
			status_msg = "Command '" + user_command + "' ran succesfully"
		else:
			status_msg = "Command '"+ user_command + "' returned exit status " + str(cmd_status)

		return status_msg, func_stat
	
	#write value to excel file, writes into first empty cell under specified column	
	def write_excel(self, args):
		func_stat = 2
		excel_value = args.split(",", 1)[0]
		file_path = args.split(",", 2)[1]
		sheet_name = args.split(",", 3)[2]
		column = args.split(",", 4)[3]

		regex = re.compile("^[A-Z]{1,3}$")
		if not regex.match(column.upper()):
			error_msg = "Invalid column '" + column + "'"
			return error_msg, func_stat

		if not os.path.exists(file_path):
			error_msg = "File '" + file_path + "' does not exist"
			return error_msg, func_stat

		if excel_value == "$serial":
			excel_value = self.get_serial()
		elif excel_value == "$mac":
			excel_value = self.get_mac()
		elif excel_value == "$date":
			excel_value = datetime.datetime.now().strftime ("%d/%m/%Y")

		wb = openpyxl.load_workbook(filename = file_path)

		if not sheet_name in wb.sheetnames:
			error_msg = "Sheet '" + sheet_name + "' does not exist"
			return error_msg, func_stat

		ws = wb[sheet_name]

		count = 1

		while 1:
			total = column.upper() + str(count)
			valuecheck = ws[total].value
			if not valuecheck:
				ws[total] = excel_value
				break
			count += 1

		if ws[total].value == excel_value:
			status_msg = "Value '" + excel_value + "' succesfully updated to " + file_path
			func_stat = 0
		else:
			status_msg =  "Failed to update value '" + excel_value + "' to " + file_path

		wb.save(filename = file_path)

		return status_msg, func_stat

#	def write_csv(self, args):
#		csv_value = args.split(",", 1)[0]
#		file_path = args.split(",", 2)[1]
#
#		if not os.path.exists(file_path):
#			error_msg = "File '" + file_path + "' does not exist"
#			return error_msg
#
#		if csv_value == "$serial":
#			csv_value = get_serial(self.conn)
#		elif csv_value == "$mac":
#			csv_value = get_mac(self.conn)

	# gets router's serial number
	def get_serial(self):
		cmd = "status -v sys |grep \"Serial Number\" |awk '{print $4}'"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)
		outp = ssh_stdout.readlines()
		serial = outp[0].strip()
		return serial

	# gets router's primary mac address
	def get_mac(self):
		cmd = "ifconfig eth0 |grep \"HWaddr\" |awk '{print $5}'"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)
		outp = ssh_stdout.readlines()
		mac = outp[0].strip()
		return mac

	# updates router's single configuration parameter, those parameters are in /etc/settings* files
	def sws(self, args):
		func_stat = 2
		param = args.split("=", 1)[0] + "="
		val = args.split("=", 2)[1]
		get_remote_file = "for i in /etc/settings.*;do grep -l '" + param  + "' $i;done"
		try:
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(get_remote_file)
			cmd_output = ssh_stdout.readlines()
			remote_file = cmd_output[0].strip()
		except IndexError:
			error_msg = "Invalid parameter '" + param + "'"
			return error_msg, func_stat
		except Exception as e:
			return e, func_stat

		if val == "$mac":
			val = self.get_mac()
		elif val == "$serial":
			val = self.get_serial()

		#change_val = "sed -i 's/" + param + ".*/" + param + val + "/' " + remote_file
		change_val = "sed -i 's|" + param + ".*|" + param + val + "|' " + remote_file
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(change_val)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		if cmd_status == 0:
			status_msg = "Parameter '" + param + "' succesfully changed to '" + val + "'"
			func_stat = 0
		else:
			status_msg = "Failed to change parameter '" + param + "'  to '" + val + "'"

		return status_msg, func_stat

	# deletes given user
	def delete_user(self, args):
		func_stat = 2
		username = args.split(",", 1)[0]
		del_cmd = "deluser " + username

		check_user = "id -u " + username
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_user)
		cmd_status = ssh_stdout.channel.recv_exit_status()
		if cmd_status != 0:
			error_msg = "User '" + username + "' doesn't exist"
			func_stat = 1
			return error_msg, func_stat

		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(del_cmd)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		if cmd_status == 0:
			status_msg = "User '" + username + "' deleted"
			func_stat = 0
		else:
			status_msg = "Failed to delete user '" + username + "'"
		return status_msg, func_stat

	# reboots router
	def reboot(self):
		cmd = "reboot"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)

	# this function creates a new user, which can be administrator or regular user
	def create_user(self, args):
		func_stat = 2
		username = args.split(",", 1)[0]
		password = args.split(",", 2)[1]
		admin = args.split(",", 3)[2]

		if "($serial)" in password:
			serial = self.get_serial()
			password = password.replace("($serial)", serial)

		if "($mac)" in password:
			mac = self.get_mac()
			password = password.replace("($mac)", mac)

		# check if user already exists
		check_user = "id -u " + username
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_user)
		cmd_status = ssh_stdout.channel.recv_exit_status()
		if cmd_status == 0:
			error_msg = "User '" + username + "' already exists"
			func_stat = 1
			return error_msg, func_stat

		# check if new user will be regular user or admin and create new user accordingly
		if admin == "admin":
			cmd = "adduser -D -S " + username + " -G root"
		else:
			cmd = "adduser -D " + username + " " + username
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		# if user was successfully created, set password for the new user
		if cmd_status == 0:
			cmd = "echo " + "'" + username + ":" + password + "' |chpasswd -m"
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)
			cmd_status = ssh_stdout.channel.recv_exit_status()
			if cmd_status == 0:
				status_msg = "New user '" + username + "' created"
				func_stat = 0
			else:
				status_msg = "New user '" + username + "' was created, but password remains blank"
				func_stat = 1
		else:
			status_msg = "Couldn't create new user " + username

		return status_msg, func_stat

	# this function changes password of given user
	def change_passwd(self, args):
		func_stat = 2
		username = args.split(",", 1)[0]
		password = args.split(",", 2)[1]

		if "($serial)" in password:
			serial = self.get_serial()
			password = password.replace("($serial)", serial)

		if "($mac)" in password:
			mac = self.get_mac()
			password = password.replace("($mac)", mac)

		# check if given user exists
		check_user = "id -u " + username
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_user)
		cmd_status = ssh_stdout.channel.recv_exit_status()
		if cmd_status != 0:
			error_msg = "Can't change password, user '" + username + "' does not exist"
			func_stat = 1
			return error_msg, func_stat

		# change password for given user
		cmd = "echo " + "'" + username + ":" + password + "' |chpasswd -m"
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(cmd)
		cmd_status = ssh_stdout.channel.recv_exit_status()
		if cmd_status == 0:
			status_msg = "Password for user '" + username + "' succesfully changed"
			func_stat = 0
		else:
			status_msg = "Failed to change password for user '" + username + "'"

		return status_msg, func_stat

	def remove_um(self, args):
		func_stat = 2
		um_name = args.split(",", 1)[0]
		um_path = "/opt/" + um_name

		# check if given user module exists in router
		check_um_path = "ls " + um_path
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_um_path)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		if cmd_status != 0:
			error_msg = "User module '" + um_name + "' doesn't exist... double check user module's name"
			return error_msg, func_stat

		# remove user module from router
		rm_um = "rm -rf " + um_path
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(rm_um)
		cmd_status = ssh_stdout.channel.recv_exit_status()

		if cmd_status !=0:
			error_msg = "Failed to remove user module '" + um_name + "'"
			return error_msg, func_stat
		else:
			status_msg = "User module '" + um_name + "' successfully removed from router"
			func_stat = 0
			return status_msg, func_stat

	# this function installs specified user module to router and verifies that md5sums match
	def add_um(self, args):
		func_stat = 2
		path_to_um = args.split(",", 1)[0]

		# check if user module exists locally
		if not os.path.exists(path_to_um):
			error_msg = "User module '" + path_to_um + "' does not exist"
			return error_msg, func_stat

		#check file's MIME type, and determine if it's in the valid format
		file_type_cmd = "file -b --mime-type " + path_to_um
		file_type = subprocess.check_output([file_type_cmd], shell=True).decode('utf-8').strip()

		if file_type != "application/gzip":
			error_msg = "Invalid user module detected (" + path_to_um + "), supported format is application/gzip"
			return error_msg, func_stat

		# check user module's name
		um_name_cmd = "tar -tf " + path_to_um + " |sed -e 's@/.*@@' |uniq"
		um_name = subprocess.check_output([um_name_cmd], shell=True).decode('utf-8').strip()

		# store local user module's md5sum
		local_md5 = LocalHash.calculate_md5(path_to_um)
		#local_cmd = "openssl md5 " + path_to_um + " |awk '{print $2}'"
		#local_md5 = subprocess.check_output([local_cmd], shell=True).decode('utf-8').strip()

		remote_path = "/root/" + os.path.basename(path_to_um)
		remote_md5_cmd = "openssl md5 " + remote_path + " |awk '{print $2}'"
		rm_remote_tar = "rm " + remote_path

		init_set = "cp /opt/" + um_name + "/etc/defaults /opt/" + um_name + "/etc/settings"
		extract_cmd = "tar -xzf " + remote_path + " -C /opt/"

		self.sftp.put(path_to_um, remote_path)

		# store remote user module's md5sum
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(remote_md5_cmd)
		cmd_output = ssh_stdout.readlines()
		remote_md5 = cmd_output[0].strip()

		# if md5sums match = user module was transferred successfully
		if local_md5 == remote_md5:

			# extract and install user module in router
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(extract_cmd)
			cmd_status = ssh_stdout.channel.recv_exit_status()
			if cmd_status == 0:
				self.conn.exec_command(init_set)
				self.conn.exec_command(rm_remote_tar)
				status_msg = "User module " + um_name + " succesfully installed to router!"
				func_stat = 0
			else:
				status_msg = "User module " + um_name + " transferred to router, but failed to install"
		else:
			status_msg = "User module " + um_name + " transferred to router, but failed integrity check, not installing..."

		return status_msg, func_stat
		

	# this function transfers configuration file to router and verifies that md5sums match
	# by default makes it active configuration (standard), but another profile can be given
	def upload_cfg(self, args):
		func_stat = 2
		unique = False
		path_to_cfg = args.split(",", 1)[0]
		profile = args.split(",", 2)[1]
		profiles = ["standard", "alt1", "alt2", "alt3"]

		# check profile validity
		if profile not in profiles:
			error_msg = "Invalid profile '" + profile +"'"
			return error_msg, func_stat

		# if $unique is given as a path, program will look for router's unique config template
		# under /etc/cossh/configs, and use it if found
		if path_to_cfg == "$unique":
			serial = self.get_serial()
			path_to_cfg = "/etc/cossh/configs/cossh_" + str(serial) + ".cfg"
			unique = True

		# check if configuration file given exists			
		if not os.path.exists(path_to_cfg) and unique == False:
			error_msg = "Invalid path to configuration file '" + path_to_cfg + "'"
			return error_msg, func_stat

		# if unique configuration file is given, check if it is in place
		elif not os.path.exists(path_to_cfg) and unique == True:
			error_msg = "Missing unique configuration file '" + path_to_cfg + "'"
			return error_msg, func_stat

		# store local configuration file's md5sum
		#cmd = "openssl md5 " + path_to_cfg + " |awk '{print $2}'"
		#local_md5 = subprocess.check_output([cmd], shell=True).decode('utf-8').strip()
		local_md5 = LocalHash.calculate_md5(path_to_cfg)

		remote_path = "/root/" + os.path.basename(path_to_cfg)
		remote_md5_cmd = "openssl md5 " + remote_path + " |awk '{print $2}'"
		restore_cfg = "restore " + remote_path

		self.sftp.put(path_to_cfg, remote_path)

		# store remote configuration file's md5sum
		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(remote_md5_cmd)
		cmd_output = ssh_stdout.readlines()
		remote_md5 = cmd_output[0].strip()

		# if md5sums match = configuration file was transferred successfully
		if local_md5 == remote_md5:

			# grep remote file for PROFILE=
			# if PROFILE= exists, override it
			# else insert PROFILE in the first line
			check_prof = "grep -q 'PROFILE=' " + remote_path
			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(check_prof)
			check_status = ssh_stdout.channel.recv_exit_status()
			if check_status == 0:
				if profile == "standard":
					add_profile = "sed -i 's/PROFILE=.*/PROFILE=/' " + remote_path
				else:
					add_profile = "sed -i 's/PROFILE=.*/PROFILE=" + profile + "/' " + remote_path
			else:
				if profile == "standard":
					add_profile = "sed -i '1 i\PROFILE=' " + remote_path
				else:
					add_profile = "sed -i '1 i\PROFILE=" + profile + "' " + remote_path

			ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(add_profile)
			cmd_status = ssh_stdout.channel.recv_exit_status()

			# restore configuration
			if cmd_status == 0:
				ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(restore_cfg)
				cmd_status = ssh_stdout.channel.recv_exit_status()
				remove_file = "rm -f " + remote_path
				ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(remove_file)
				if cmd_status == 0:
					status_msg = "Configuration file " + os.path.basename(path_to_cfg)  + " uploaded to " + profile
					func_stat = 0
				else:
					status_msg = "Failed to upload configuration file to " + profile
			else:
				status_msg = "Couldn't activate profile '" + profile + "', configuration failed"
		else:
			status_msg = "File " + path_to_cfg + " was transferred to router, but failed integrity check, can't activate it"
	
		return status_msg, func_stat
	

	# uploads given file to given location in router
	def upload_file(self, args):
		func_stat = 2
		path_in_orig = args.split(",", 1)[0]
		path_in_dest = args.split(",", 2)[1]
		filename = os.path.basename(path_in_orig)
		check_remote_path = os.path.basename(path_in_dest)

		if not os.path.exists(path_in_orig):
			error_msg = "File '" + path_in_orig + "' does not exist"
			return error_msg, func_stat

		if filename != check_remote_path:
			if path_in_dest.endswith("/"):
				path_in_dest = path_in_dest + filename
			else:
				path_in_dest = path_in_dest + "/" + filename

		#cmd = "openssl md5 " + path_in_orig + " |awk '{print $2}'"
		#local_md5 = subprocess.check_output([cmd], shell=True).decode('utf-8').strip()
		local_md5 = LocalHash.calculate_md5(path_in_orig)
		remote_md5_cmd = "openssl md5 " + path_in_dest + " |awk '{print $2}'"

		try:
			self.sftp.put(path_in_orig, path_in_dest)
		except FileNotFoundError:
			error_msg = "Invalid local or remote path noticed while trying to transfer file " + filename
			return error_msg, func_stat

		ssh_stdin, ssh_stdout, ssh_stderr = self.conn.exec_command(remote_md5_cmd)
		cmd_output = ssh_stdout.readlines()
		remote_md5 = cmd_output[0].strip()

		if local_md5 == remote_md5:
			status_msg = "File '" + filename + "' succesfully transferred to router"
			func_stat = 0
		else:
			status_msg = "File '" + filename + "' couldn't be transferred, or it failed integrity check"

		return status_msg, func_stat

